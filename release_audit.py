#!/usr/bin/env python3
"""Audit the final human-facing release artifacts."""
from __future__ import annotations

import json
from pathlib import Path
import subprocess
import zipfile

from openpyxl import load_workbook
from PIL import Image


ROOT = Path(__file__).resolve().parents[1]
INTEGRATED = ROOT / "integrated"
FIGURES = INTEGRATED / "figures"
MANUSCRIPT = INTEGRATED / "manuscript"
TABLES = INTEGRATED / "tables"
QC = INTEGRATED / "qc"


def load(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def pdf_pages(path: Path) -> int:
    result = subprocess.run(
        ["pdfinfo", str(path)], check=True, text=True,
        stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
    )
    for line in result.stdout.splitlines():
        if line.startswith("Pages:"):
            return int(line.split(":", 1)[1].strip())
    raise RuntimeError(f"No page count in pdfinfo output for {path}")


def valid_zip(path: Path) -> bool:
    with zipfile.ZipFile(path) as archive:
        return archive.testzip() is None


def main() -> None:
    checks: list[dict[str, object]] = []

    def record(name: str, passed: bool, detail: object = "") -> None:
        checks.append({"check": name, "pass": bool(passed), "detail": str(detail)})

    main_png = sorted(FIGURES.glob("Figure_*.png"))
    supplement_png = sorted(FIGURES.glob("Supplementary_Figure_*.png"))
    record("eight main PNG figures", len(main_png) == 8, len(main_png))
    record("four supplementary PNG figures", len(supplement_png) == 4, len(supplement_png))
    for path in main_png + supplement_png:
        raw = path.read_bytes()
        with Image.open(path) as image:
            image.verify()
        record(f"valid PNG {path.name}", raw.startswith(b"\x89PNG\r\n\x1a\n") and raw.endswith(b"IEND\xaeB`\x82"), len(raw))
        record(f"PDF peer {path.stem}", path.with_suffix(".pdf").exists())
        record(f"SVG peer {path.stem}", path.with_suffix(".svg").exists())
    record("no partial figure files", not any(FIGURES.glob(".*.tmp.*")))

    report_docx = MANUSCRIPT / "INH_Integrated_Final_Analysis_Report.docx"
    article_docx = MANUSCRIPT / "INH_Revised_Article_With_Integrated_Analysis.docx"
    report_pdf = report_docx.with_suffix(".pdf")
    article_pdf = article_docx.with_suffix(".pdf")
    for path in (report_docx, article_docx):
        record(f"valid Office archive {path.name}", valid_zip(path), path.stat().st_size)
    record("report PDF has 17 pages", pdf_pages(report_pdf) == 17, pdf_pages(report_pdf))
    record("article PDF has 43 pages", pdf_pages(article_pdf) == 43, pdf_pages(article_pdf))

    for stem in ("INH_Integrated_Final_Analysis_Report", "INH_Revised_Article_With_Integrated_Analysis"):
        audit = load(QC / f"{stem}_a11y_final.json")
        counts = audit["counts"]
        record(f"zero accessibility findings: {stem}", sum(counts.values()) == 0, counts)

    workbook_path = TABLES / "INH_Integrated_Source_Data.xlsx"
    record("valid XLSX archive", valid_zip(workbook_path), workbook_path.stat().st_size)
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    expected_sheets = {
        "README", "Study_Design", "Stage2_Effects", "Stage3_Pairwise",
        "Stage3_Ablations", "Stage4_Validation", "Stage4_Stress",
        "Stage5_Scenarios", "Stage5_Boundaries",
    }
    record("all source-data sheets present", set(workbook.sheetnames) == expected_sheets, workbook.sheetnames)
    record("workbook cross-sheet formulas retained", workbook["README"]["B7"].value == "=B8" and workbook["README"]["B8"].value == "=COUNTA('Stage4_Validation'!F2:F7)")
    workbook.close()

    embedded = load(QC / "embedded_test_report.json")
    record("embedded tests", embedded["failed_count"] == 0 and embedded["passed_count"] == 51, embedded)
    independent = load(QC / "independent_recomputation.json")
    record("141 independent recomputations", independent["all_checks_pass"] and len(independent["checks"]) == 141, len(independent["checks"]))
    quality = load(QC / "quality_gate.json")
    record("59-item quality gate", quality["passed"] and quality["passed_count"] == quality["total_count"] == 59, quality["passed_count"])

    seed_rows = (QC / "stage3_seed_robustness.csv").read_text(encoding="utf-8").strip().splitlines()
    record("six Stage 3 independent-root reruns", len(seed_rows) == 7, len(seed_rows) - 1)

    article_text = "\n".join(
        paragraph.text for paragraph in __import__("docx").Document(article_docx).paragraphs
    )
    record("article discloses missing editorial metadata", "Submission Metadata to Complete" in article_text)
    record("article retains limited Stage 5 decision", "Stage 5 therefore remained a limited sensitivity analysis" in article_text)

    output = {
        "release": "INH integrated v1.0.0",
        "checks": checks,
        "passed": all(item["pass"] for item in checks),
        "passed_count": sum(item["pass"] for item in checks),
        "total_count": len(checks),
    }
    QC.mkdir(parents=True, exist_ok=True)
    (QC / "release_audit.json").write_text(json.dumps(output, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({key: output[key] for key in ("passed", "passed_count", "total_count")}, indent=2))
    if not output["passed"]:
        raise SystemExit(2)


if __name__ == "__main__":
    main()

